#プログラム1｜ライブラリ設定
from datetime import datetime #excelファイル名作成
import openpyxl as px #Excelファイルの操作
from openpyxl.styles import PatternFill #Excelの背景色変更
import requests #Web情報の取得
from bs4 import BeautifulSoup #ウェブスクレイピングでデータを抽出
import main
#プログラム2｜ヤフオクページから情報をスクレイピング
def Pagecrawling(soup, item_list):

    # プログラム2-1｜<li>タグのなかで、class='Product'のものを変数productsに格納
    products = soup.find_all('li', class_='Product') #liタグのProductのクラスから、各製品情報を取得｜プログラム2-1

    
#プログラム2-1からタグとクラスで分解（製品の詳細情報取得のための下準備）｜プログラム2-2~2-5
    # プログラム2-2｜変数productsの要素を一つずつ調査
    for product in products:
        # プログラム2-3｜変数productsのなかで、aタグで、class='Product__titleLink'のものを変数urlsに格納
        urls = product.find_all('a', class_='Product__titleLink')

        # プログラム2-4｜変数productsのなかで、divタグで、class='Product__priceInfo'のものを変数pricesに格納
        prices = product.find_all('div', class_='Product__priceInfo')

        # プログラム2-5｜変数productsのなかで、divタグで、class='Product__otherInfo'のものを変数othersに格納
        others = product.find_all('div', class_='Product__otherInfo')
#製品の詳細情報を取得｜プログラム2-6~2-17
        # プログラム2-6｜zip関数でまとめて繰り返し処理を実行
        for (url, price, other) in zip(urls, prices, others):

            # プログラム2-7｜変数linkに変数urlのhref部分を取得（製品url）
            link = url.get('href')

            # プログラム2-8｜変数nameに変数urlのtitle部分を取得（製品名）
            name = url.get('title')

            # プログラム2-9｜変数priceのspanタグで、class='Product__price'のものを変数pricevaluesに格納
            pricevalues = price.find_all('span', class_='Product__price')

            # プログラム2-10｜変数の初期化
            currentprice = '-'
            fixedprice = '-'

            # プログラム2-11｜変数pricevaluesの要素を一つずつ処理
            for pricevalue in pricevalues:

                # プログラム2-12｜もし変数pricevalueのテキスト情報が'現在'が含まれていれば
                if '現在' in pricevalue.get_text():

                    # プログラム2-13｜変数currentpriceに現在の価格を取得（'\n'で改行を削除）
                    currentprice = pricevalue.get_text().replace('\n', '')

                # プログラム2-14｜もし変数pricevalueのテキスト情報が'即決'が含まれていれば
                elif '即決' in pricevalue.get_text():

                    # プログラム2-15｜変数fixedpriceに即決の価格を取得（'\n'で改行を削除）
                    fixedprice = pricevalue.get_text().replace('\n', '')

            # プログラム2-16｜変数labelのdivタグの0番目の要素を取得（'\n'で改行を削除）→入札
            label = other.find_all('div')[0].get_text().replace('\n','')

            # プログラム2-17｜変数labelのdivタグの1番目の要素を取得（'\n'で改行を削除）→残り時間
            lefttime = other.find_all('div')[1].get_text().replace('\n','')

            # プログラム2-18｜リスト「item_list」に必要な要素を追加
            item_list.append([name, link, currentprice, fixedprice, label, lefttime]) #item_list」のりストへ製品の詳細情報を追加｜プログラム2-18

    # プログラム2-19｜リスト「item_list」を返す
    return item_list #「item_list」をプログラム5へ返す

# プログラム3｜次へリンクを探す
def Checkpage(pagerlinks):

    # プログラム3-1｜変数pageurlをNoneにリセット
    pageurl = None

    # プログラム3-2｜引数pagerlinksの中に、「次へ」があれば、そのurlをpageurlに設定（「次へ」がない場合は、Noneのまま）
    for pagerlink in pagerlinks:
        if pagerlink.get_text() == '次へ':
            pageurl = pagerlink.get('href')
            break

    # プログラム3-3｜pageurlを返す
    return  pageurl

# プログラム4｜エクセルに出力
def Write_excel(item_list, keyword):

    # プログラム4-1｜エクセルを取得
    wb = px.Workbook()
    ws = wb.active

    # プログラム4-2｜エクセルのヘッダーの背景色を設定
    fill = PatternFill(patternType='solid', fgColor='e0e0e0', bgColor='e0e0e0')

    # プログラム4-3｜エクセル1行目のヘッダーを出力
    headers = ['No', 'タイトル', '現在の価格','即決価格','入札','残り時間']
    for i, header in enumerate(headers):
        ws.cell(row=1, column=1+i, value=headers[i])
        ws.cell(row=1, column=1+i).fill = fill

    # プログラム4-4｜エクセル2行目以降のデータを出力
    for y, row in enumerate(item_list):
        ws.cell(row= y+2, column= 1, value= y+1)
        for x, cell in enumerate(row):
            if x == 0:
                ws.cell(row= y+2, column= x+2, value=item_list[y][x])
            elif x == 1:
                ws.cell(row= y+2, column= x+1).hyperlink = item_list[y][x]
                ws.cell(row= y+2, column= x+1).font = px.styles.fonts.Font(color='0000EE')
            else:
                ws.cell(row= y+2, column= x+1, value=item_list[y][x])

    # プログラム4-5｜日付を取得
    now = datetime.now()
    hiduke = now.strftime('%Y-%m-%d')

    # プログラム4-6｜エクセルファイルの保存
    filename = hiduke + '_' + keyword  + '_' +'YahooAuction.xlsx'
    wb.save(filename)

#プログラム5｜mainプログラム
# プログラム5-1｜検索キーワードとYahooオークションURLの設定
keyword = input('出力したい商品名')
url = 'https://auctions.yahoo.co.jp/search/search?p=' + keyword + '&n=100';

# プログラム5-2｜ヤフオクページの取得
r = requests.get(url)
soup = BeautifulSoup(r.text, 'lxml')

# プログラム5-3｜取得したデータを格納するリスト
item_list = []

# プログラム5-4｜リスト「item_list」にプログラム3の結果を格納
item_list = Pagecrawling(soup, item_list)

# プログラム5-5｜プログラム4を実行して、次のページがあるかどうか調査
pageurl = Checkpage(soup.find_all('a', class_='Pager__link'))

# プログラム5-6｜次のページがなくなるまで、処理を実行
while True:
    # プログラム5-7｜次のページがないならwhile文を終了
    if pageurl is None:
        break

    # プログラム5-8｜次のページがあるなら処理を実行
    else:
        # プログラム5-9｜次のページを取得
        nextpage = requests.get(pageurl)
        soup = BeautifulSoup(nextpage.text, 'lxml')

        # プログラム5-10｜リスト「item_list」にプログラム3の結果を格納
        item_list = Pagecrawling(soup, item_list)

        # プログラム5-11｜プログラム4を実行して、次のページがあるかどうか調査
        pageurl = Checkpage(soup.find_all('a', class_='Pager__link'))

# プログラム5-12｜リスト「item_list」の中身を調べる
[print(i, item) for i, item in enumerate(item_list)]

# プログラム5-13｜プログラム5を実行して、リスト「item_list」をエクセルに書き出す
Write_excel(item_list, keyword)